"""
Excel file parser for importing existing data.
"""
import openpyxl
from datetime import datetime
from pathlib import Path
from backend.database import get_db


async def import_excel(file_path: str) -> dict:
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if "ACTIVITYLOG" not in wb.sheetnames:
        raise ValueError("Excel file must contain 'ACTIVITYLOG' sheet")

    ws = wb["ACTIVITYLOG"]
    stats = {"imported": 0, "skipped": 0, "errors": []}

    db = await get_db()
    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                entry_id, timestamp, activity, entry_type, follow_up, status, reference_id, eta = row

                if not activity:
                    stats["skipped"] += 1
                    continue

                entry_type = entry_type or "LOG"
                status = status or "LOG"

                if isinstance(timestamp, datetime):
                    timestamp_str = timestamp.isoformat()
                else:
                    timestamp_str = str(timestamp) if timestamp else datetime.now().isoformat()

                eta_str = eta.isoformat() if isinstance(eta, datetime) else (str(eta) if eta else None)

                await db.execute(
                    """INSERT INTO log_entries (timestamp, activity, type, follow_up, status, reference_id, eta, created_at, updated_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        timestamp_str,
                        activity,
                        entry_type,
                        follow_up,
                        status,
                        reference_id,
                        eta_str,
                        datetime.now().isoformat(),
                        datetime.now().isoformat()
                    )
                )
                stats["imported"] += 1
            except Exception as e:
                stats["errors"].append(f"Row {row_idx}: {str(e)}")

        await db.commit()
        return stats
    finally:
        await db.close()


async def export_excel(file_path: str) -> dict:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ACTIVITYLOG"

    headers = ["ID", "timestamp", "activity", "type", "Follow up", "STATUS", "REFERENCE ID", "ETA"]
    ws.append(headers)

    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT * FROM log_entries ORDER BY timestamp DESC"
        )
        rows = await cursor.fetchall()

        for row in rows:
            entry = dict(row)
            ws.append([
                entry["id"],
                entry["timestamp"],
                entry["activity"],
                entry["type"],
                entry.get("follow_up"),
                entry.get("status"),
                entry.get("reference_id"),
                entry.get("eta")
            ])

        wb.save(file_path)
        return {"exported": len(rows), "file": file_path}
    finally:
        await db.close()
